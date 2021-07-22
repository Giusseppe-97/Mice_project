"""
Created on Fri Apr  2 15:18:14 2021

@author: josep
"""

# Executing the program as a HD window for windows and exception for running it on mac
try:
    from ctypes import windll
    windll.shcore.SetProcessDpiAwareness(1)
except:
    pass

import tkinter as tk
from tkinter import ttk
from tkcalendar import Calendar as tkc
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import askdirectory

from matplotlib.backends.backend_tkagg import (
    FigureCanvasTkAgg, NavigationToolbar2Tk)
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator

import pandas as pd
import seaborn as sns

import os
from datetime import datetime as dt

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import nicexcel as nl

print("Loading TheMiceCounter application. This might take a minute.")

class Application(tk.Tk):
    """[Creating a main App class where all the frames are going to be set upon]

    The main App class must inherit from tk.Tk which is the root or main window
    """

    def __init__(self, *args, **kwargs):
        """init is the method for setting default state of the object
        """

        super().__init__(*args, **kwargs)


        # Set style of the GUI
        self.tk.call('source', r'packages/style_azure/azure.tcl')
        ttk.Style().theme_use('azure')

        # Create Main Frames
        self.mainFrame1 = tk.Frame(self)
        self.mainFrame2 = tk.Frame(self)

        # Create Calendar
        self.cal = tkc(
            self.mainFrame1, selectbackground="#120597", background="#120597",
            selectmode="day", year=2021, month=5, day=1
        )

        # Call methods
        self.configure_basic_tk_properties()
        self.pack_all()

    def configure_basic_tk_properties(self):
        """This method configures the basic tkinter esthetic properties for the GUI
        """
        self.title("  TheMiceCounter")

        # Setting the main App in the center regardless to the window's size chosen by the user
        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)

        # Setting a background for the main App which will be divided in 2 different Frames
        self.configure(bg="light blue")

        # Creating and placing Lables for each Frame
        self.lable = tk.Label(
            self.mainFrame1, text="RUN DATA", foreground="white",
            background="#120597").place(x=0, width=1920
                                        )
        self.lable2 = tk.Label(
            self.mainFrame2, text="HISTOGRAM DISPLAY PREVIEW", foreground="white",
            background="#120597").place(y=0, width=1920
                                        )

        # Creating labels
        self.label3 = ttk.Label(
            self.mainFrame1, text="Input: ", background="white"
        )
        self.label4 = ttk.Label(
            self.mainFrame1, text="Output:", background="white"
        )
        self.label5 = tk.Label(
            self.mainFrame1, text="Select Date Interval: ", background="white"
        )
        self.label6 = ttk.Label(
            self.mainFrame1, text="From:", background="white"
        )
        self.label7 = ttk.Label(
            self.mainFrame1, text="To:", background="white"
        )

        # Creating textboxes
        self.textbox1 = ttk.Entry(self.mainFrame1, width=80)
        self.textbox2 = ttk.Entry(self.mainFrame1, width=80)
        self.textbox3 = ttk.Entry(self.mainFrame1, width=20)
        self.textbox4 = ttk.Entry(self.mainFrame1, width=20)

        # Creating and initializing buttons
        self.button1 = ttk.Button(
            self.mainFrame1, text="Select", command=lambda: [self.open_excel_file_location()]
        )
        self.button2 = ttk.Button(
            self.mainFrame1, text="Save", command=lambda: [self.save_results()]
        )
        self.button3 = ttk.Checkbutton(
            self.mainFrame1, text="Run", style='ToggleButton', command=self.display_plot
        )
        self.button4 = ttk.Button(
            self, text="End date", command=self.grab_end_date
        )
        self.button5 = ttk.Button(
            self, text="Start date", command=self.grab_start_date
        )
        self.button_reset = ttk.Button(
            master=self, text="Reset", command=self.reset_app
        )
        self.button_quit = ttk.Button(
            master=self, text="Quit", command=self.quit
        )

        # Creating Canvas (where Histograms are going to be placed as matplotlib Figures)
        self.canvas01 = tk.Canvas(self.mainFrame2)
        self.canvas02 = tk.Canvas(self.mainFrame2)

    def grab_start_date(self):
        if len(self.textbox3.get()) != 0:
            self.textbox3.delete(0, 'end')
        self.textbox3.insert(tk.END, self.cal.selection_get())
        self.init_month = self.cal.selection_get().strftime("%B")

    def grab_end_date(self):
        if len(self.textbox4.get()) != 0:
            self.textbox4.delete(0, 'end')
        self.textbox4.insert(tk.END, self.cal.selection_get())
        self.final_month = self.cal.selection_get().strftime("%B")

    def pack_all(self):

        self.mainFrame1.place(x=0, y=0, height=600, width=1950)
        self.mainFrame2.place(x=0, y=200, rely=0.05, height=1000, width=1950)

        self.label3.place(x=20, y=70, height=40, width=80)
        self.label4.place(x=20, y=140, height=40, width=80)
        self.label5.place(x=1300, y=30, height=30)
        self.label6.place(x=1250, y=80, height=30)
        self.label7.place(x=1250, y=150, height=30)

        self.textbox1.place(x=80, y=70, height=40, width=800)
        self.textbox2.place(x=80, y=140, height=40, width=800)
        self.textbox3.place(x=1300, y=70, height=40, width=99)
        self.textbox4.place(x=1300, y=140, height=40, width=99)

        self.button1.place(x=900, y=70, height=40, width=120)
        self.button2.place(x=900, y=140, height=40, width=120)
        self.button3.place(x=1090, y=70, height=50, width=120)
        self.button_reset.place(x=1090, y=140, height=40, width=120)
        self.button5.place(x=1450, y=70, height=40)
        self.button4.place(x=1450, y=140, height=40)
        self.button_quit.pack(side=tk.BOTTOM, pady=10)
        # quit button not placed yet, just packed
        self.cal.place(x=1550, y=30, rely=0.005, relx=0.02, height=210, width=300)

        self.canvas01.place(x=100, y=40, height=600, width=800)
        self.canvas02.place(x=1000, y=40, height=600, width=800)

    def open_excel_file_location(self):
        """Open the File Explorer to select desired excel file
        """
        global filepath1
        if len(self.textbox1.get()) != 0:
            self.textbox1.delete(0, 'end')
            
        filepath1 = askopenfilename(filetypes=[(
            "xlsx Files", "*.xlsx"), ("csv Files", "*.csv"), ("All Files", "*.*")])
        try:
            with open(filepath1, "r"):
                self.textbox1.insert(tk.END, filepath1)
        
        except:
            if not filepath1:
                tk.messagebox.showwarning(title='No file selected.',
                    message='Please make sure a file has been chosen before running the program.')
                filepath1 = "../data/R403Q SoftMouse Export.xlsx"
                with open(filepath1, "r"):
                    self.textbox1.insert(tk.END, filepath1)

    def save_results(self):
        """Open the file Explorer to select desired location to save results
        """
        global filepath2
        if len(self.textbox2.get()) != 0:
            self.textbox2.delete(0, 'end')
        self.filepath2 = askdirectory()
        if not self.filepath2:
            tk.messagebox.showwarning(title='No folder selected',
                message='Please make sure a folder has been chosen before running the program.')
            self.filepath2 = "results/2021_monthly_results/plots_per_month"
        self.textbox2.insert(tk.END, self.filepath2)

    def import_excel_file(self):
        self.df = pd.read_excel(filepath1, sheet_name="Mouse List")

    def obtain_data_from_excel(self):
        self.import_excel_file()

        # Converting date format to compare with excel dates
        self.init_date = self.textbox3.get()
        self.init_date[::-1]
        self.final_date = self.textbox4.get()
        self.final_date[::-1]

        # Limit data from the excel file for the chosen period of time
        n_df = self.df[(self.init_date <= self.df.Date_of_birth) &
                       (self.df.Date_of_birth <= self.final_date)]

        # print(n_df)

        # Failed attempt to make the code underneath better

        # lage = []
        # for index, row in n_df.iterrows():
        #     a1 = str(self.final_date)
        #     aa = dt.strptime(a1, "%Y-%m-%d")
        #     b1 = str(n_df['Date_of_birth'])
        #     b1 = b1[5:15]
        #     print(b1)
        #     bb = dt.strptime(b1, "%Y-%m-%d")
        #     bd = abs((bb - aa).days)
        #     age_in_weeks = bd//7
        #     lage.append(age_in_weeks)
        # print(lage)

        # List of ages of the different type of mice 
        birth_MWt = []
        birth_FWt = []
        birth_MHet = []
        birth_FHet = []

# ***
# My BEST idea so far is to reorganize this code. If I calculate the birthday of all of the dataframe
# I can then just grab the dataframe of 'sex', 'Genotype' and 'status' and they will have the birthdays already
# ***

# Needs optimization. Too many lines for a simple part of the code
# There are 4 loops that are doing the same thing for example. There should be a way to reduce this
        # Selects data from the excel file for sex, genotype and status (only mice that are alive)
        df_MWt = pd.DataFrame(n_df.loc[(self.df['Sex'] == 'Male') & (
            n_df['Genotype'] == 'Null(-)') & (n_df['Status'] == 'Alive')])

        for index, rows in df_MWt.iterrows():
            a = str(n_df['Date_of_birth'][index])
            birth_MWt.append(a[:10])

        df_FWt = pd.DataFrame(n_df.loc[(self.df['Sex'] == 'Female') & (
            n_df['Genotype'] == 'Null(-)') & (n_df['Status'] == 'Alive')])

        for index, rows in df_FWt.iterrows():
            a = str(n_df['Date_of_birth'][index])
            birth_FWt.append(a[:10])

        df_MHet = pd.DataFrame(n_df.loc[(self.df['Sex'] == 'Male') & (
            n_df['Genotype'] == 'R403Q(+/-)') & (n_df['Status'] == 'Alive')])

        for index, rows in df_MHet.iterrows():
            a = str(n_df['Date_of_birth'][index])
            birth_MHet.append(a[:10])

        df_FHet = pd.DataFrame(n_df.loc[(self.df['Sex'] == 'Female') & (
            n_df['Genotype'] == 'R403Q(+/-)') & (n_df['Status'] == 'Alive')])

        for index, rows in df_FHet.iterrows():
            a = str(n_df['Date_of_birth'][index])
            birth_FHet.append(a[:10])

        dfappended = df_FHet.append([df_FWt,df_MHet, df_MWt], ignore_index=True)
        self.dfreduced = dfappended[['Sex', 'Genotype', 'Status','Date_of_birth']]

        birthday = [ birth_FHet, birth_FWt, birth_MHet, birth_MWt]

        # Using the dates entered by the user to calculate de age of the mice
        self.d2 = []
        dd2 = []
        
        for i in range(len(birthday)):
            age_in_weeks = 0
            d1 = []
            j = 0
            for j in range(len(birthday[i])):

                a1 = str(self.final_date)
                aa = dt.strptime(a1, "%Y-%m-%d")
                b1 = str(birthday[i][j])
                bb = dt.strptime(b1, "%Y-%m-%d")
                bd = abs((bb - aa).days)
                age_in_weeks = bd//7

                dd2.append(age_in_weeks)
                d1.append(age_in_weeks)
                j = j + 1
            d1.sort(reverse=True)
            self.d2.append(d1)

        # Adding a column in excel with the calculated age
        self.dfreduced.loc[:,'Calculated Age'] = dd2

        # Colors
        self.colors = sns.color_palette("rocket", 4)

        # Setting parameters of length and size for the matplotlib plots
        self.d3 = []
        self.d4 = []
        es = 0
        r = 0
        self.listy = []
        for es in range(4):
            if self.d2[es] != []:
                self.d3.append(self.d2[es][0])
            else:
                self.d4.append(es)
        if len(self.d4) == 4:
            self.b = 5
        else:
            self.b = int(max(self.d3))+1
        for i in range(self.b+1):
            self.listy.append(r)
            r = r+1

    def plot_individual_hist(self):

        self.obtain_data_from_excel()

        # Plot
        fig = plt.figure(figsize=(5, 5), dpi=100)

        f = fig.gca()
        f.hist(self.d2, bins=self.listy, color=self.colors)
        f.set_xlabel(r'Age (weeks)', fontsize=15)
        f.set_ylabel(r'Number of mice', fontsize=17)
        f.set_title('NUMBER OF MICE VS AGE',
                    horizontalalignment='center', fontweight="bold", fontsize=20)
        f.xaxis.set_major_locator(MaxNLocator(nbins=len(self.listy)))
        f.xaxis.set_major_locator(MaxNLocator(integer=True))

        f.legend(['Male Null(-)', 'Female Null(-)', 'Male R403Q(+/-)', 'Female R403Q(+/-)'])

        return fig

    def plot_4_hist(self):

        self.obtain_data_from_excel()

        # Plot
        fig, axs = plt.subplots(2, 2, figsize=(
            8, 8), sharey=True,  tight_layout=True)

        # Defining histogram characteristics guaranteeing the same x axis quantity and integers
        axs[0, 0].hist(self.d2[0], bins=self.listy, color=self.colors[0],
                       label='Male Null(-)', edgecolor='white')

        axs[0, 1].hist(self.d2[1], bins=self.listy, color=self.colors[1],
                       label='Female Null(-)', edgecolor='white')

        axs[1, 0].hist(self.d2[2], bins=self.listy, color=self.colors[2],
                       label='Male R403Q(+/-)', edgecolor='white')

        axs[1, 1].hist(self.d2[3], bins=self.listy, color=self.colors[3],
                       label='Female R403Q(+/-)', edgecolor='white')

    # Unverified but presumably better way to write the lines underneath
        # for i in range(1):
        #     for j in range(1):
        #         axs[i, j].xaxis.set_major_locator(MaxNLocator(integer=True))
        # for i in range(1):
        #     for j in range(1):
        #         axs[i, j].yaxis.set_major_locator(MaxNLocator(integer=True))

        axs[0, 0].xaxis.set_major_locator(MaxNLocator(integer=True))
        axs[0, 1].xaxis.set_major_locator(MaxNLocator(integer=True))
        axs[1, 0].xaxis.set_major_locator(MaxNLocator(integer=True))
        axs[1, 1].xaxis.set_major_locator(MaxNLocator(integer=True))
        axs[0, 0].yaxis.set_major_locator(MaxNLocator(integer=True))
        axs[0, 1].yaxis.set_major_locator(MaxNLocator(integer=True))
        axs[1, 0].yaxis.set_major_locator(MaxNLocator(integer=True))
        axs[1, 1].yaxis.set_major_locator(MaxNLocator(integer=True))

        axs[1, 0].set_xlabel(r'Age (weeks)', fontsize=15)
        axs[1, 1].set_xlabel(r'Age (weeks)', fontsize=15)
        axs[0, 0].set_ylabel(r'Number of mice', fontsize=17,
                             horizontalalignment='right')
        axs[0, 0].legend()
        axs[0, 1].legend()
        axs[1, 0].legend()
        axs[1, 1].legend()

        plt.suptitle('NUMBER OF MICE VS AGE',
                     horizontalalignment='center', fontweight="bold", fontsize=20)

        return fig, axs

    def display_plot(self):
        """plot function is created for plotting the graph in tkinter window
        """
        # creating two Tkinter canvas that will contain the Matplotlib figures
        fig = self.plot_individual_hist()
        fig2, ax = self.plot_4_hist()

        # creating canvas and drawing figures into canvas
        self.canvas2 = FigureCanvasTkAgg(fig, master=self.canvas02)
        self.canvas1 = FigureCanvasTkAgg(fig2, master=self.canvas01)
        self.canvas1.draw()
        self.canvas2.draw()

        # creating the Matplotlib toolbars
        self.toolbar1 = NavigationToolbar2Tk(
            self.canvas1, self, pack_toolbar=False)
        self.toolbar2 = NavigationToolbar2Tk(
            self.canvas2, self, pack_toolbar=False)
        self.toolbar1.update()
        self.toolbar2.update()

        # placing the canvas on the Tkinter window
        self.toolbar1.place(x=400, y=900)
        self.toolbar2.place(x=1300, y=900)

        # packing the widgets inside the Tkinter canvas
        self.canvas1.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        self.canvas2.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)

        # Calling the function that creates the excel sheets

        self.create_excel_file()

    def load_workbook(path_workbook_data):
        print("LOADING...")
        """
        Loading an excel workbook that already exists
        or opening a new one if it does not

        Parameters
        ----------
        path_workbook_data : [str]
            [path to the excel workbook location]
        """
        if os.path.exists(path_workbook_data):
            print('old file') 
            return openpyxl.load_workbook(path_workbook_data)
        print('new file') 
        return openpyxl.Workbook()

    def create_excel_file(self):
        
        self.fm = FolderManager(self)
        path_workbook_data = "results/2021_monthly_results/data_results.xlsx"
        plot_name_fig_four = str(self.fm.plot_4_name)
        path_image = "results/2021_monthly_results/plots_per_month/" + plot_name_fig_four + ".png"

        wb = load_workbook(path_workbook_data)
        wb.create_sheet(title=self.final_date,index = 0)
        sheet = wb[self.final_date]
        rows = dataframe_to_rows(self.dfreduced,index=False)
        
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20

        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

            
        img = openpyxl.drawing.image.Image(path_image)
        img.height = 400
        img.width = 400
        sheet.add_image(img, 'H1')

        for column in range(1, sheet.max_column +1):
            cell = sheet.cell(row=1, column=column)
            cell.style = 'Pandas'

        wb.save(path_workbook_data)
        wb.close()

    def reset_app(self):
        self.canvas1.get_tk_widget().destroy()
        self.canvas2.get_tk_widget().destroy()
        self.toolbar1.destroy()
        self.toolbar2.destroy()

class FolderManager:
    """
    Enables the developement and management of folder structures for multipurpose 
    projects with datetime configurations.
    :param project_name: name of the project for the FolderManager agent.
    """

    def __init__(self, app_object):
        self.app_object = app_object
        self.current_datetime = dt.now()
        self.get_current_important_values()
        self.generate_folder_paths()
        self.create_folders()
        self.save_image()

    def get_current_important_values(self):
        self.year = self.current_datetime.strftime("%Y")
        self.month = self.current_datetime.strftime("%B")
        self.week = self.current_datetime.strftime("%W")

    def save_image(self):
        if self.app_object.init_month != self.app_object.final_month:
            self.plot_4_name = str(self.app_object.init_month) + \
                "-" + str(self.app_object.final_month)
        else:
            self.plot_4_name = str(self.app_object.init_month)

        self.filepath_4_plot = str(self.app_object.filepath2) + "/" + \
            str(self.plot_4_name)+".png"
        self.hist_4_plot = plt.savefig(self.filepath_4_plot)

    def generate_folder_paths(self):
        self.current_directory = os.path.abspath(os.path.dirname(__file__))

        self.directory_per_month = os.path.join(
            self.current_directory,
            "results",
            "{}_{}".format(
                self.year,
                "monthly_results"
            )
        )
        self.directory_plots_month = os.path.join(
            "results",
            "2021_monthly_results",
            "{}".format("plots_per_month")
        )

    def create_folders(self):
        if not os.path.exists(self.directory_per_month):
            os.makedirs(self.directory_per_month)

        if not os.path.exists(self.directory_plots_month):
            os.makedirs(self.directory_plots_month)

    def get_path_for_results(self):
        return self.directory_per_month


if __name__ == "__main__":
    app = Application()
    app.geometry("1900x990+0+0")
    app.resizable(True, False)
    app.iconbitmap(r'../docs/_site/favicon.ico')
    app.mainloop()