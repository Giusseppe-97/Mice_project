
from combined_gui import FolderManager 
#Libreria (necesario descargarla) para la creacion de los archivos de Excel
import xlsxwriter
import  os
import pandas as pd
from openpyxl import load_workbook
from xlsxwriter.utility import xl_rowcol_to_cell
import nicexcel as nl


# I need a function to create and saves excel sheets with the gui info
def crear_excel_de_proyectos(self):
    # From folderManager I have a folder where to save the info
    self.fm = FolderManager(self)
    path_workbook_data = "results/2021_monthly_results/data_results.xlsx"
    plot_name_four = str(self.fm.plot_4_name)

    # Se utiliza la libreria de creacion de excels para crear un archivo xlsx con todos los proyectos y su info
    # if path and excel object exists: 
    if os.path.isfile(path_workbook_data):

        print('old file')
        op_writer = pd.ExcelWriter(path_workbook_data, engine = 'openpyxl')
        op_writer.book = load_workbook(path_workbook_data)
        op_writer.book.create_sheet("{}".format(plot_name_four))
        
        nl.to_excel(self.dfreduced, path_workbook_data, sheet_name=self.fm.plot_4_name, index = False)
        op_writer.save()
        op_writer.close()

    else:

        self.wb = xlsxwriter.Workbook( "{}\\{}".format( self.fm.directory_per_month,'{}.xlsx'.format("data_results")  ) )

