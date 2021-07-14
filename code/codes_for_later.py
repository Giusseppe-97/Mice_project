from openpyxl import Workbook
import datetime
import os

import xlsxwriter
path_workbook_data = "results/2021_monthly_results/data_results.xlsx"
sheet_time = str(datetime.datetime.now().strftime("%S"))


workbook  = xlsxwriter.Workbook('filename.xlsx')
if os.path.isfile(path_workbook_data):

        worksheet = workbook.add_worksheet(sheet_time)
        worksheet.write(0, 0, 'Hello Excel_1')

workbook.close()

wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 35

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
ws['A3'] = datetime.datetime.now()



# for sheet_time in wb:
#         ws1 = wb.create_sheet(sheet_time) 

#         if ws1 == sheet_time.title:
#                 # insert at the end (default)
#                 ws1 = wb.create_sheet(sheet_time)
                



# # Save the file
# wb.save("sample.xlsx")










#     #Libreria para manejo de la fecha y el guardado de informacion segun la fecha
# import datetime
# #Libreria para manejar la lectura correcta de los archivos TXT
# import glob
# #Libreria (necesario descargarla) para la creacion de los archivos de Excel
# import xlsxwriter
# from xlsxwriter.utility import xl_rowcol_to_cell


# def crear_excel_de_proyectos(self):
# #Se crea la carpeta donde se guarda la informacion de los proyectos
# path_carpeta = INFO_DIR_ACTUAL.CREACION_CARPETA()
# #Se utiliza la libreria de creacion de excels para crear un archivo xlsx con todos los proyectos y su info
# PROYECTOS_EXCEL = xlsxwriter.Workbook( "{}\\{}".format( path_carpeta.devolver_ruta_guardado() ,'{}.xlsx'.format("PROYECTOS")  ) )

# #Formatos utiles para mostrar informacion organizada
# bold = PROYECTOS_EXCEL.add_format({'bold': True})
# bold_border = PROYECTOS_EXCEL.add_format({"bold":True,"border":True})
# bold_border_gris = PROYECTOS_EXCEL.add_format({"bold":True,"border":True,"align":"center","bg_color":"#C8C8C8"})
# normal_border = PROYECTOS_EXCEL.add_format({"border":True})
# currency_border = PROYECTOS_EXCEL.add_format({"num_format":'$#,##0.00',"border":True})
# currency_border_bold  = PROYECTOS_EXCEL.add_format({"num_format":'$#,##0.00',"border":True,"bold":True})
# txt_rojo_bold = PROYECTOS_EXCEL.add_format({'bold': True, 'font_color':"red"})
# txt_verde_bold = PROYECTOS_EXCEL.add_format({"bold":True, "font_color":"green"})
# fondo_gris_bold = PROYECTOS_EXCEL.add_format({"bold":True,"bg_color":"#C0C0C0","border":True})
# fondo_azul_bold_merge_1 = PROYECTOS_EXCEL.add_format({"bold":True,"bg_color":"#66B2FF","align":"center","border":True})
# fondo_azul_bold_merge_2 = PROYECTOS_EXCEL.add_format({"bold":True,"bg_color":"#CFE2F5","align":"left","border":True,"font_color":"black"})



# #Creamos vector respectivas hojas de excel para cada uno de los proyectos (una por cada proyecto)
# vector_hojas_excel = []

# #Accedemos a cada proyecto...
# for proy in range( len(self.vector_proyectos) ):
#         #Creamos variables encargadas del manejo de filas y columnas en el archivo de excel para cada hoja de cada proyecto
#         fila = 5
#         columna = 0 

#         #Agregamos tantos proyectos como sean necesarios en un vector asociados a las hojas de excel (con su nombre)
#         vector_hojas_excel.append( PROYECTOS_EXCEL.add_worksheet("{}".format( self.vector_proyectos[proy].nombre_proyecto ) )  )
        


#         #Dejamos el nombre del proyecto respectivo en la primera celda, junto con la info basica
#         vector_hojas_excel[proy].write( 0,0, "{}".format( "NOMBRE PROYECTO:" ), fondo_gris_bold )
#         vector_hojas_excel[proy].write( 0,1, "{}".format( self.vector_proyectos[proy].nombre_proyecto ), fondo_gris_bold )
#         vector_hojas_excel[proy].write( 1,0, "{}".format( "FECHA INICIO:" ), fondo_gris_bold )
#         vector_hojas_excel[proy].write( 1,1, "{}".format( self.vector_proyectos[proy].fecha_inicio_proyecto ), fondo_gris_bold)
#         vector_hojas_excel[proy].write( 2,0, "{}".format( "FECHA FIN:" ), fondo_gris_bold )
#         vector_hojas_excel[proy].write( 2,1, "{}".format( self.vector_proyectos[proy].fecha_fin_proyecto),fondo_gris_bold )
#         vector_hojas_excel[proy].write( 3,0, "{}".format( "CANTIDAD ETAPAS:" ), fondo_gris_bold )
#         try:
#         vector_hojas_excel[proy].write( 3,1, int(self.vector_proyectos[proy].cantidad_etapas), fondo_gris_bold )
#         except:
#         vector_hojas_excel[proy].write( 3,1,self.vector_proyectos[proy].cantidad_etapas, fondo_gris_bold )




#         #Accedemos a cada etapa...
#         for etap in range( len( self.vector_proyectos[proy].vector_etapas) ):
#         #Cada etapa tendra las filas de un tamanno horizontal optimo (mejorar visualizacion y separacion entre ellas) 
#         vector_hojas_excel[proy].set_column('{}:{}'.format( xl_rowcol_to_cell(fila,columna) , xl_rowcol_to_cell(fila,columna + 3) ), 30)
        
#         #Mostramos titulo de cada etapa del proyecto
#         vector_hojas_excel[proy].merge_range( fila,columna,fila,columna + 3, "{}".format( str(self.vector_proyectos[proy].vector_etapas[etap].nombre_etapa)  ), fondo_azul_bold_merge_1 )
#         fila = fila + 1
        
#         #Mostramos personal de trabajo de la etapa (director, coordinador y arquitectos)
#         vector_hojas_excel[proy].write( fila, columna, "DIRECTOR ETAPA:",bold )
#         vector_hojas_excel[proy].write( fila, columna + 1 , "{}".format( self.vector_proyectos[proy].vector_etapas[etap].director ),bold )
#         vector_hojas_excel[proy].write( fila + 1, columna, "COORDINADOR ETAPA:",bold )
#         vector_hojas_excel[proy].write( fila + 1, columna + 1 , "{}".format( self.vector_proyectos[proy].vector_etapas[etap].coordinador ),bold )
#         fila = fila + 2

#         #Mostramos arquitectos, segun los que hayan...
#         if len( self.vector_proyectos[proy].vector_etapas[etap].vector_arquitectos ) == 0:
#                 vector_hojas_excel[proy].write( fila, columna, "ARQUITECTOS:",bold )
#                 vector_hojas_excel[proy].write( fila, columna + 1, "PENDIENTE POR ASIGNAR",bold )
#                 fila = fila + 1

#         else:
#                 vector_hojas_excel[proy].write( fila, columna, "ARQUITECTOS:",bold )
#                 for arq in range( len( self.vector_proyectos[proy].vector_etapas[etap].vector_arquitectos ) ):
#                 vector_hojas_excel[proy].write( fila, columna + 1, "{}".format( self.vector_proyectos[proy].vector_etapas[etap].vector_arquitectos[arq] ) ,bold )
#                 fila = fila + 1
#         vector_hojas_excel[proy].write( fila + 1, columna,"ÁREA ETAPA:",bold )
#         try:
#                 vector_hojas_excel[proy].write( fila + 1, columna + 1,float( self.vector_proyectos[proy].vector_etapas[etap].area_etapa)  ,bold )
#         except:
#                 vector_hojas_excel[proy].write( fila + 1, columna + 1,self.vector_proyectos[proy].vector_etapas[etap].area_etapa,bold )

#         fila = fila + 4
#         #Recorremos cada subetapa (siempre son las mismas 4)
#         for subetapa in range(4):
#                 #Con esto comenzamos a mostrar la info de cada subetapa (su nombre)
#                 vector_hojas_excel[proy].merge_range( fila,columna, fila, columna + 3,"{}".format( self.vector_proyectos[proy].vector_etapas[etap].vector_subetapas[subetapa].nombre_subetapa ),fondo_azul_bold_merge_2 )                    
#                 fila = fila +1
#                 vector_hojas_excel[proy].write( fila,columna,"{}".format( "ESTADO ACTUAL" ),bold )
#                 vector_hojas_excel[proy].write( fila,columna + 1,"{}".format( self.vector_proyectos[proy].vector_etapas[etap].vector_subetapas[subetapa].estado_subetapa ),bold )
#                 fila = fila +1
#                 vector_hojas_excel[proy].write( fila,columna,"{}".format( "FECHA INICIO" ),bold )
#                 vector_hojas_excel[proy].write( fila,columna + 1,"{}".format( self.vector_proyectos[proy].vector_etapas[etap].vector_subetapas[subetapa].fecha_inicio ),bold )
#                 fila = fila +1
#                 vector_hojas_excel[proy].write( fila,columna,"{}".format( "FECHA FINALIZACIÓN" ),bold )
#                 vector_hojas_excel[proy].write( fila,columna + 1,"{}".format( self.vector_proyectos[proy].vector_etapas[etap].vector_subetapas[subetapa].fecha_fin ),bold )
#                 fila = fila +3
                
#                 #Comenzamos a mostrar la info de los gastos
#                 vector_hojas_excel[proy].write( fila,columna,"{}".format( "GASTOS:" ),txt_rojo_bold )
#                 fila = fila +1

#                 #Condicional para mostrar gastos (o en su defecto, mostrar que no hay)
#                 if len( self.vector_proyectos[proy].vector_etapas[etap].vector_subetapas[subetapa].gastos ) == 0:
#                 #Mostramos que no hay gastos, si el vector de gastos esta vacio
#                 vector_hojas_excel[proy].write( fila,columna,"{}".format( "No hay gastos registrados" ) )
#                 fila = fila +1
#                 else:
#                 #Si SI hay gastos, se muestra el encabezado para la fecha, el nombre del gasto y el valor asociado
#                 vector_hojas_excel[proy].write( fila,columna,"{}".format("Fecha"),bold_border_gris)
#                 vector_hojas_excel[proy].write( fila,columna + 1,"{}".format("Nombre Gasto"),bold_border_gris)
#                 vector_hojas_excel[proy].write( fila,columna + 2,"{}".format("Cantidad"),bold_border_gris)
#                 fila = fila + 1

#                 #Se crea string con las CELDAS que ALMACENAN LA CANTIDAD DE LOS GASTOS (para ingresar formula del total)
#                 string_celdas_gastos = "="
#                 #Se recorren los gastos, con ayuda del vector que los almacena y se muestra su info importante
#                 for gasto in range( len(self.vector_proyectos[proy].vector_etapas[etap].vector_subetapas[subetapa].gastos) ):
#                         vector_hojas_excel[proy].write( fila, columna,"{}".format( self.vector_proyectos[proy].vector_etapas[etap].vector_subetapas[subetapa].gastos[gasto].fecha_gasto ), normal_border )
#                         vector_hojas_excel[proy].write( fila, columna + 1,"{}".format( self.vector_proyectos[proy].vector_etapas[etap].vector_subetapas[subetapa].gastos[gasto].nombre_gasto ),normal_border )
#                         vector_hojas_excel[proy].write( fila, columna + 2, float(self.vector_proyectos[proy].vector_etapas[etap].vector_subetapas[subetapa].gastos[gasto].cantidad_gasto ),currency_border )

#                         #Se almacenan las celdas de los gastos estrategicos
#                         if gasto == ( len(self.vector_proyectos[proy].vector_etapas[etap].vector_subetapas[subetapa].gastos) - 1 ):
#                         string_celdas_gastos = string_celdas_gastos + xl_rowcol_to_cell(fila,columna+2)
#                         else:
#                         string_celdas_gastos = string_celdas_gastos + xl_rowcol_to_cell(fila,columna+2) + "+"

#                         #Se aumenta la fila para continuar debido proceso
#                         fila = fila + 1


#                 # # print(string_celdas_gastos)
#                 vector_hojas_excel[proy].write( fila,columna + 1,"{}".format("total gastos") , bold_border )
#                 vector_hojas_excel[proy].write( fila,columna + 2, string_celdas_gastos, currency_border_bold)
#                 fila = fila + 1


#                 #Ahora se muestran los Ingresos
#                 fila = fila + 1
#                 vector_hojas_excel[proy].write( fila,columna,"{}".format( "INGRESOS:" ),txt_verde_bold )
#                 fila = fila + 1
#                 #Condicional para mostrar ingresos (o en su defecto, mostrar que NO hay)
#                 if len( self.vector_proyectos[proy].vector_etapas[etap].vector_subetapas[subetapa].ingresos ) == 0:
#                 #Mostramos que no hay ingresos, si el vector de ingresos esta vacio
#                 vector_hojas_excel[proy].write( fila,columna,"{}".format( "No hay ingresos registrados" ) )
#                 fila = fila +1
#                 else:
#                 #Si SI hay ingresos, se muestra el encabezado para la fecha, el nombre del ingreso y el valor asociado
#                 vector_hojas_excel[proy].write( fila,columna,"{}".format("Fecha"),bold_border_gris)
#                 vector_hojas_excel[proy].write( fila,columna + 1,"{}".format("Nombre Ingreso"),bold_border_gris)
#                 vector_hojas_excel[proy].write( fila,columna + 2,"{}".format("Cantidad"),bold_border_gris)
#                 fila = fila + 1

#                 #Se recorren los ingresos, con ayuda del vector que los almacena y se muestra su info importante
#                 #OJO: se debe crear un string que logre almacenar las celdas que llevan cantidad de ingreso (para formula posterior)
#                 string_celdas_ingresos = "="
#                 for ingreso in range( len(self.vector_proyectos[proy].vector_etapas[etap].vector_subetapas[subetapa].ingresos) ):
#                         vector_hojas_excel[proy].write( fila, columna,"{}".format( self.vector_proyectos[proy].vector_etapas[etap].vector_subetapas[subetapa].ingresos[ingreso].fecha_ingreso ), normal_border )
#                         vector_hojas_excel[proy].write( fila, columna + 1,"{}".format( self.vector_proyectos[proy].vector_etapas[etap].vector_subetapas[subetapa].ingresos[ingreso].nombre_ingreso ),normal_border )
#                         vector_hojas_excel[proy].write( fila, columna + 2, float(self.vector_proyectos[proy].vector_etapas[etap].vector_subetapas[subetapa].ingresos[ingreso].cantidad_ingreso ),currency_border )

#                         #Se almacenan las celdas de los gastos estrategicos
#                         if ingreso == ( len(self.vector_proyectos[proy].vector_etapas[etap].vector_subetapas[subetapa].ingresos) - 1 ):
#                         string_celdas_ingresos = string_celdas_ingresos + xl_rowcol_to_cell(fila,columna+2)
#                         else:
#                         string_celdas_ingresos = string_celdas_ingresos + xl_rowcol_to_cell(fila,columna+2) + "+"

#                         #Se aumenta la fila para continuar debido proceso
#                         fila = fila + 1


#                 # # print(string_celdas_gastos)
#                 vector_hojas_excel[proy].write( fila,columna + 1,"{}".format("total ingresos") , bold_border )
#                 vector_hojas_excel[proy].write( fila,columna + 2, string_celdas_ingresos, currency_border_bold)
#                 fila = fila + 1


#                 #Dejamos un espacio de 3 entre cada subetapa (para ser mas organizados)
#                 fila = fila + 3



#         columna = columna + 5
#         fila = 5

        
                        
# PROYECTOS_EXCEL.close()

# d2 = [[6 ,6 ,6 ,5 ,5, 2, 2, 2, 2, 1, 0], [4, 3, 3, 2, 2, 1, 1, 1, 1, 0, 0, 0, 0], [0], [3, 1, 0]]
# months_2021_excel_wb = Workbook()
# worksheet_month_ws =  months_2021_excel_wb.active
# worksheet_month_ws.title = "Changed Sheet"

# longest_list_size = 0
# list_dic = []
# i = 0
# j = 0
# k = 0
# l = 0

# for k in range(4):
#         if len(d2[k])>len(d2[0]):
#                 longest_list_size = len(d2[k]) 

# dic_mic={}
# for longest_list_size in range(longest_list_size,-1,-1):
#         dic_mic.setdefault(longest_list_size,[]).append(0)

# list_of_dic = []
# data = ['Male Wildtype', 'Female Wildtype', 'Male Heterozygous', 'Female Heterozygous']

# for i in range(4):
        
#         dic_mice = dict()
#         num_of_mice = []
#         age_mice_weeks = []
#         for j in range(len(d2[i])):

#                 if d2[i][j] not in age_mice_weeks:
#                         age_mice_weeks.append(d2[i][j])
#                         num_of_mice = d2[i].count(d2[i][j])
#                         dic_mice.setdefault(d2[i][j],[]).append(num_of_mice)

#         dic_mice = {key: dic_mice.get(key, dic_mic[key]) for key in dic_mic}
#         df = pd.DataFrame.from_dict(dic_mice) 
#         list_dic.append(df)
        
#         print(data[i])
#         print(df.to_string(index=False))
# print(list_dic)

# result = pd.concat(list_dic)
# print(result.to_string( index = False))

# rows = dataframe_to_rows(result,index = False)

# for r_idx, row in enumerate(rows, 1):
#         for c_idx, value in enumerate(row, 1):
#                 worksheet_month_ws.cell(row=r_idx, column=c_idx, value=value)

# months_2021_excel_wb.save(filename = 'results/sample_book.xlsx')
 


# birthdays = []
#         i = 0
#         j = 0
#         for index, rows in total_data[i].iterrows():
#             aa = str(total_data[i]['Date_of_birth'])
#             print(aa[0:9:-1])
#             bb = str(self.final_date)
#             # print(bb)
            
#             birthdays.append(aa)

#         print(birthdays)

# dt = pd.DataFrame({
#             'Male Wildtype': d2[0]},{
#                 'Female Wildtype': d2[1]},{
#                     'Male Heterozygous': d2[2]},{
#                         'Female Heterozygous': d2[3]})

# print(dt)

# self.pop_up_messages()

# # Create scrollbar
# self.scroll_frame_1 = ttk.Scrollbar(self.mainFrame1, orient="vertical", command=self.canvas1.yview)
# self.scroll_frame_1.pack(side='right', fill='y')




# def pop_up_messages(self):
#     if str(self.textbox1) == "" and str(self.textbox3) != "":
#         messagebox.showinfo("Warning Message: Empty input filepath",
#                             "Please select a valid Excel file or write its path before continuing")

#     if str(self.textbox2) == "" and str(self.textbox3) != "" & str(self.textbox4) != "":
#         messagebox.showinfo("Warning Message: Empty output filepath",
#                             "Please select a valid folder")

#     if str(self.textbox1)[-1] != "x":
#         messagebox.showinfo("Warning Message",
#                             "Please select a valid Excel file path before continuing")

#     if str(self.textbox3)[3] != "1" or str(self.textbox4)[3] != "1":
#         messagebox.showinfo("Warning Message: Invalid date or date period",
#                             "Please select a valid date (from 2021 further)")


# self.directory_per_week = os.path.join(
        #     self.current_directory,
        #     "results",
        #     "{}_{}".format(
        #         self.year,
        #         "weekly_results"
        #     )
        # )
        # self.directory_plots_week = os.path.join(
        #     "results",
        #     "2021_weekly_results",
        #     "{}".format("plots_per_week")
        # )

        # if not os.path.exists(self.directory_per_week):
        #     os.makedirs(self.directory_per_week)
        # if not os.path.exists(self.directory_plots_week):
        #     os.makedirs(self.directory_plots_week)




# def create_excel_workbook_weeks(self):
        # Create excel workbooks (xlsx)

        # self.weeks_2021_excel = xlsxwriter.Workbook("{}\\{}".format(
        #     self.directory_per_week, '{}.xlsx'.format("Weeks_2021")))

        # i = 1
        # self.excel_sheet_vector_week = []
        # for i in range(53):
        #     self.excel_sheet_vector_week.append("week " + str(i))

        # # Esthetic paramenters for Week's excel
        # normal_text = self.weeks_2021_excel.add_format(
        #     {'bold': False, "align": "center"}
        # )
        # bold = self.weeks_2021_excel.add_format(
        #     {'bold': True, "align": "center"}
        # )
        # bold_border = self.weeks_2021_excel.add_format(
        #     {"bold": True, "align": "center", "border": True}
        # )

        # self.excel_sheet_vector_week.append(self.weeks_2021_excel.add_worksheet(
        #     "{}".format(self.excel_sheet_vector_week[0])))

        # self.weeks_2021_excel.close()







        # worksheet_month.write('A2', mg_o.df_MWt)
        # worksheet_month.write('D2', mg_o.df_FWt)
        # worksheet_month.write('G2', mg_o.df_MHet)
        # worksheet_month.write('J2', mg_o.df_FHet)
        # filep = str(FigureManager.generate_figures)
        # worksheet_month.insert_image('D14', filep )

        # # Esthetic paramenters for Month's excel
        # normal_text = self.months_2021_excel.add_format(
        #     {'bold': False, "align": "center"}
        #     )
        # bold = self.months_2021_excel.add_format(
        #     {'bold': True, "align": "center"}
        #     )
        # bold_border = self.months_2021_excel.add_format(
        #     {"bold": True, "align": "center", "border": True}
        #     )

        # # time datetime for loop to add sheets in positions of the vector
        # self.excel_sheet_vector_month.append(self.months_2021_excel.add_worksheet(
        #         "{}".format(self.excel_sheet_vector_month[0][1])))



        # months_2021_excel_wb = Workbook()
        # worksheet_month_ws = months_2021_excel_wb.active
        # worksheet_month_ws.title = "Changed Sheet"

        # longest_list_size = len(d2[0])
        # list_dic = []
        # i = 0
        # j = 0
        # k = 0
        # l = 0

        # for k in range(4):
        #     if len(d2[k]) > len(d2[0]):
        #         longest_list_size = len(d2[k])
            
        # dic_mic = {}
        # for longest_list_size in range(longest_list_size, -1, -1):
        #     dic_mic.setdefault(longest_list_size, []).append(0)

        # list_of_dic = []
        # data = ['Male Wildtype', 'Female Wildtype',
        #         'Male Heterozygous', 'Female Heterozygous']

        # for i in range(4):

        #     dic_mice = dict()
        #     num_of_mice = []
        #     age_mice_weeks = []
        #     for j in range(len(d2[i])):

        #         if d2[i][j] not in age_mice_weeks:
        #             age_mice_weeks.append(d2[i][j])
        #             num_of_mice = d2[i].count(d2[i][j])
        #             dic_mice.setdefault(d2[i][j], []).append(num_of_mice)

        #     dic_mice = {key: dic_mice.get(
        #         key, dic_mic[key]) for key in dic_mic}
        #     df = pd.DataFrame.from_dict(dic_mice)
        #     list_dic.append(df)

        #     print(data[i])
        #     print(df.to_string(index=False))
        # # print(list_dic)

        # result = pd.concat(list_dic)
        # # print(result.to_string(index=False))

        # rows = dataframe_to_rows(result, index=False)

        # for r_idx, row in enumerate(rows, 1):
        #     for c_idx, value in enumerate(row, 1):
        #         worksheet_month_ws.cell(row=r_idx, column=c_idx, value=value)

        # months_2021_excel_wb.save(filename='results/sample_book.xlsx')

               # self.hist_4_plot.anchor(self.plot_name.cell('A20'))
        # self.plot_name.add_image(self.hist_4_plot)
        # months_2021_excel_wb.save("{}\\{}".format(
        #     self.directory_per_month, '{}.xlsx'.format("Months_2021")))
